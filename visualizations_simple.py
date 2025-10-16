import psycopg2
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime
import numpy as np
 
# Configuration
TIMELINE_PLAYER_LIMIT = 24

# Database connection parameters
DB_CONFIG = {
    'host': 'localhost',
    'database': 'data_v',
    'user': 'postgres',
    'password': '0412',
    'port': '5432'
}

def connect_to_db():
    """Establish database connection"""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        return conn
    except Exception as e:
        print(f"Error connecting to database: {e}")
        return None

def execute_query(conn, query, description):
    """Execute a query and return DataFrame"""
    try:
        df = pd.read_sql_query(query, conn)
        return df
    except Exception as e:
        print(f"Error executing query: {e}")
        return None

def create_charts_directory():
    """Create charts directory if it doesn't exist"""
    if not os.path.exists('charts'):
        os.makedirs('charts')
    if not os.path.exists('exports'):
        os.makedirs('exports')

def create_pie_chart(conn):
    """Create pie chart: Distribution of teams by average rating"""
    query = """
    SELECT 
        ps.team,
        ROUND(AVG(ps.rating), 2) as avg_rating,
        COUNT(*) as player_count
    FROM player_stats ps
    GROUP BY ps.team
    ORDER BY avg_rating DESC
    """
    
    df = execute_query(conn, query, "Team Average Ratings")
    if df is None or df.empty:
        return 0
    
    # Create pie chart
    plt.figure(figsize=(12, 8))
    colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', 
              '#DDA0DD', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E9']
    
    wedges, texts, autotexts = plt.pie(df['avg_rating'], 
                                      labels=df['team'], 
                                      autopct='%1.1f%%',
                                      colors=colors[:len(df)],
                                      startangle=90)
    
    plt.title('Distribution of Teams by Average Rating\n(Valorant Champions 2024)', 
              fontsize=16, fontweight='bold', pad=20)
    plt.axis('equal')
    
    # Add legend with team names and ratings
    legend_labels = [f"{team}: {rating}" for team, rating in zip(df['team'], df['avg_rating'])]
    plt.legend(wedges, legend_labels, title="Teams (Avg Rating)", 
               loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    plt.tight_layout()
    plt.savefig('charts/team_rating_distribution_pie.png', dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"Pie Chart: Team rating distribution - {len(df)} teams")
    return len(df)

def create_bar_chart(conn):
    """Create bar chart: Top 10 players by ACS (Average Combat Score)"""
    query = """
    SELECT 
        ps.player_name,
        ps.team,
        ps.acs,
        ps.rating
    FROM player_stats ps
    WHERE ps.rounds > 100
    ORDER BY ps.acs DESC
    LIMIT 10
    """
    
    df = execute_query(conn, query, "Top Players by ACS")
    if df is None or df.empty:
        return 0
    
    # Create bar chart
    plt.figure(figsize=(14, 8))
    bars = plt.bar(range(len(df)), df['acs'], color='skyblue', edgecolor='navy', linewidth=1.2)
    
    # Customize bars with different colors
    colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', 
              '#DDA0DD', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E9']
    for i, bar in enumerate(bars):
        bar.set_color(colors[i % len(colors)])
    
    plt.xlabel('Players', fontsize=12, fontweight='bold')
    plt.ylabel('Average Combat Score (ACS)', fontsize=12, fontweight='bold')
    plt.title('Top 10 Players by Average Combat Score\n(Valorant Champions 2024)', 
              fontsize=16, fontweight='bold', pad=20)
    
    # Set x-axis labels
    plt.xticks(range(len(df)), df['player_name'], rotation=45, ha='right')
    
    # Add value labels on bars
    for i, (bar, acs, rating) in enumerate(zip(bars, df['acs'], df['rating'])):
        plt.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 1, 
                f'{acs:.1f}\n(R: {rating:.2f})', 
                ha='center', va='bottom', fontsize=9, fontweight='bold')
    
    plt.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    plt.savefig('charts/top_players_acs_bar.png', dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"Bar Chart: Top players by ACS - {len(df)} players")
    return len(df)

def create_horizontal_bar_chart(conn):
    """Create horizontal bar chart: Map win rates by side (Attack vs Defense)"""
    query = """
    SELECT 
        ms.map_name,
        ms.attack_win_percent,
        ms.defense_win_percent,
        ms.times_played
    FROM maps_stats ms
    ORDER BY ms.times_played DESC
    """
    
    df = execute_query(conn, query, "Map Win Rates")
    if df is None or df.empty:
        return 0
    
    # Create horizontal bar chart
    fig, ax = plt.subplots(figsize=(12, 8))
    
    y_pos = range(len(df))
    
    # Create bars for attack and defense
    bars1 = ax.barh(y_pos, df['attack_win_percent'], height=0.4, 
                    label='Attack Win %', color='lightcoral', alpha=0.8)
    bars2 = ax.barh([y + 0.4 for y in y_pos], df['defense_win_percent'], height=0.4, 
                    label='Defense Win %', color='lightblue', alpha=0.8)
    
    ax.set_yticks([y + 0.2 for y in y_pos])
    ax.set_yticklabels(df['map_name'])
    ax.set_xlabel('Win Percentage (%)', fontsize=12, fontweight='bold')
    ax.set_title('Map Win Rates: Attack vs Defense\n(Valorant Champions 2024)', 
                 fontsize=16, fontweight='bold', pad=20)
    ax.legend()
    ax.grid(axis='x', alpha=0.3)
    
    # Add value labels
    for i, (attack, defense) in enumerate(zip(df['attack_win_percent'], df['defense_win_percent'])):
        ax.text(attack + 1, i + 0.2, f'{attack:.1f}%', va='center', fontweight='bold')
        ax.text(defense + 1, i + 0.6, f'{defense:.1f}%', va='center', fontweight='bold')
    
    plt.tight_layout()
    plt.savefig('charts/map_win_rates_horizontal_bar.png', dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"Horizontal Bar Chart: Map win rates - {len(df)} maps")
    return len(df)

def create_line_chart(conn):
    """Create line chart: Player performance over time (matches)"""
    query = """
    SELECT 
        m.date,
        m.match_id,
        dmps.player_name,
        dmps.rating,
        dmps.acs,
        dmps.map_name
    FROM matches m
    INNER JOIN detailed_matches_player_stats dmps ON m.match_id = dmps.match_id
    WHERE dmps.stat_type = 'map' 
    AND dmps.player_name IN (
        SELECT player_name 
        FROM player_stats 
        WHERE rating > 1.1 
        ORDER BY rating DESC 
        LIMIT 5
    )
    ORDER BY m.date, dmps.player_name
    """
    
    df = execute_query(conn, query, "Player Performance Over Time")
    if df is None or df.empty:
        return 0
    
    # Convert date to datetime
    df['date'] = pd.to_datetime(df['date'], format='%a, %B %d, %Y')
    
    # Create line chart
    plt.figure(figsize=(14, 8))
    
    # Plot lines for each player
    colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7']
    for i, player in enumerate(df['player_name'].unique()):
        player_data = df[df['player_name'] == player].sort_values('date')
        plt.plot(player_data['date'], player_data['rating'], 
                marker='o', linewidth=2, label=player, markersize=6, color=colors[i % len(colors)])
    
    plt.xlabel('Date', fontsize=12, fontweight='bold')
    plt.ylabel('Rating', fontsize=12, fontweight='bold')
    plt.title('Top Players Performance Over Time\n(Valorant Champions 2024)', 
              fontsize=16, fontweight='bold', pad=20)
    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
    plt.grid(True, alpha=0.3)
    plt.xticks(rotation=45)
    
    plt.tight_layout()
    plt.savefig('charts/player_performance_line.png', dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"Line Chart: Player performance over time - {len(df)} data points")
    return len(df)

def create_histogram(conn):
    """Create histogram: Distribution of player ratings"""
    query = """
    SELECT 
        ps.rating,
        ps.team
    FROM player_stats ps
    WHERE ps.rounds > 50
    """
    
    df = execute_query(conn, query, "Player Ratings Distribution")
    if df is None or df.empty:
        return 0
    
    # Create histogram
    plt.figure(figsize=(12, 8))
    
    # Create histogram with different colors for different rating ranges
    n, bins, patches = plt.hist(df['rating'], bins=20, alpha=0.7, edgecolor='black', linewidth=0.5)
    
    # Color bars based on rating ranges
    for i, (patch, bin_left, bin_right) in enumerate(zip(patches, bins[:-1], bins[1:])):
        bin_center = (bin_left + bin_right) / 2
        if bin_center < 0.8:
            patch.set_facecolor('red')
        elif bin_center < 1.0:
            patch.set_facecolor('orange')
        elif bin_center < 1.2:
            patch.set_facecolor('yellow')
        else:
            patch.set_facecolor('green')
    
    plt.xlabel('Player Rating', fontsize=12, fontweight='bold')
    plt.ylabel('Number of Players', fontsize=12, fontweight='bold')
    plt.title('Distribution of Player Ratings\n(Valorant Champions 2024)', 
              fontsize=16, fontweight='bold', pad=20)
    
    # Add statistics text
    mean_rating = df['rating'].mean()
    median_rating = df['rating'].median()
    plt.axvline(mean_rating, color='red', linestyle='--', linewidth=2, label=f'Mean: {mean_rating:.2f}')
    plt.axvline(median_rating, color='blue', linestyle='--', linewidth=2, label=f'Median: {median_rating:.2f}')
    
    plt.legend()
    plt.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    plt.savefig('charts/player_ratings_histogram.png', dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"Histogram: Player ratings distribution - {len(df)} players")
    return len(df)

def create_scatter_plot(conn):
    """Create scatter plot: ACS vs Rating correlation with team colors"""
    query = """
    SELECT 
        ps.player_name,
        ps.team,
        ps.acs,
        ps.rating,
        ps.kd_ratio,
        ps.rounds
    FROM player_stats ps
    WHERE ps.rounds > 100
    """
    
    df = execute_query(conn, query, "ACS vs Rating Correlation")
    if df is None or df.empty:
        return 0
    
    # Create scatter plot
    plt.figure(figsize=(14, 10))
    
    # Get unique teams and assign colors
    teams = df['team'].unique()
    colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', 
              '#DDA0DD', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E9', 
              '#F8C471', '#82E0AA', '#F1948A', '#85C1E9', '#D7BDE2']
    team_colors = dict(zip(teams, colors[:len(teams)]))
    
    # Create scatter plot with team colors
    for team in teams:
        team_data = df[df['team'] == team]
        plt.scatter(team_data['acs'], team_data['rating'], 
                   c=team_colors[team], label=team, alpha=0.7, s=100, edgecolors='black', linewidth=0.5)
    
    plt.xlabel('Average Combat Score (ACS)', fontsize=12, fontweight='bold')
    plt.ylabel('Rating', fontsize=12, fontweight='bold')
    plt.title('ACS vs Rating Correlation by Team\n(Valorant Champions 2024)', 
              fontsize=16, fontweight='bold', pad=20)
    
    # Add trend line
    z = np.polyfit(df['acs'], df['rating'], 1)
    p = np.poly1d(z)
    plt.plot(df['acs'], p(df['acs']), "r--", alpha=0.8, linewidth=2, label='Trend Line')
    
    # Add correlation coefficient
    correlation = df['acs'].corr(df['rating'])
    plt.text(0.05, 0.95, f'Correlation: {correlation:.3f}', 
             transform=plt.gca().transAxes, fontsize=12, 
             bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
    
    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig('charts/acs_rating_scatter.png', dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"Scatter Plot: ACS vs Rating correlation - {len(df)} players")
    return len(df)

def create_time_slider_chart(conn):
    """Create interactive Plotly chart with time slider - Top Players Performance Over Time"""
    query = f"""
    WITH eligible_players AS (
        SELECT dmps.player_name
        FROM detailed_matches_player_stats dmps
        JOIN matches m ON m.match_id = dmps.match_id
        WHERE dmps.stat_type = 'map'
        GROUP BY dmps.player_name
        ORDER BY COUNT(DISTINCT m.date) DESC, COUNT(*) DESC
        LIMIT {TIMELINE_PLAYER_LIMIT}
    ), raw AS (
        SELECT 
            m.date AS date,
            m.match_id,
            m.stage,
            dmps.player_name,
            dmps.player_team AS team,
            dmps.rating,
            dmps.acs,
            dmps.k AS kills,
            dmps.d AS deaths,
            dmps.a AS assists,
            dmps.map_name
        FROM matches m
        INNER JOIN detailed_matches_player_stats dmps ON m.match_id = dmps.match_id
        WHERE dmps.stat_type = 'map'
          AND dmps.player_name IN (SELECT player_name FROM eligible_players)
    ), ranked AS (
        SELECT 
            *,
            ROW_NUMBER() OVER (
                PARTITION BY player_name, date
                ORDER BY rating DESC, acs DESC, kills DESC, match_id DESC
            ) AS rn
        FROM raw
    )
    SELECT *
    FROM ranked
    WHERE rn = 1
    ORDER BY date, rating DESC;
    """
    
    df = execute_query(conn, query, "Top Players Performance Over Time")
    if df is None or df.empty:
        # Write placeholder HTML so the file is created even if no data
        fig = go.Figure()
        fig.update_layout(
            title="Top Players Performance Over Time - No data available",
            annotations=[dict(text="No data available for timeline", x=0.5, y=0.5, showarrow=False)]
        )
        create_charts_directory()
        fig.write_html('charts/interactive_player_performance_timeline.html')
        print("[ Time Slider Chart: No data returned; wrote placeholder HTML")
        return 0
    
    # Convert date to datetime (robust), sort, and derive daily key for animation
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    df = df.dropna(subset=['date']).sort_values(['date', 'player_name'])
    df['date_display'] = df['date'].dt.strftime('%B %d, %Y')
    # Log which players are included
    included_players = sorted(df['player_name'].unique().tolist())
    print(f"[INFO] Timeline includes {len(included_players)} players (limit {TIMELINE_PLAYER_LIMIT}): {', '.join(included_players)}")
    if len(included_players) < TIMELINE_PLAYER_LIMIT:
        print("[INFO] Fewer players than limit due to data availability across dates.")
    if df.empty:
        fig = go.Figure()
        fig.update_layout(
            title="Top Players Performance Over Time - No valid dates",
            annotations=[dict(text="No valid dates after parsing", x=0.5, y=0.5, showarrow=False)]
        )
        create_charts_directory()
        fig.write_html('charts/interactive_player_performance_timeline.html')
        print("[Time Slider Chart: No valid dates after parsing; wrote placeholder HTML")
        return 0
    # Build cumulative frames so lines connect over time
    df['date_norm'] = df['date'].dt.normalize()
    unique_frames = sorted(df['date_norm'].unique())
    cumulative_frames = []
    for frame_date in unique_frames:
        frame_df = df[df['date_norm'] <= frame_date].copy()
        frame_df['frame'] = pd.to_datetime(frame_date).strftime('%Y-%m-%d')
        cumulative_frames.append(frame_df)
    if not cumulative_frames:
        fig = go.Figure()
        fig.update_layout(
            title="Top Players Performance Over Time - No frames",
            annotations=[dict(text="No timeline frames to animate", x=0.5, y=0.5, showarrow=False)]
        )
        create_charts_directory()
        fig.write_html('charts/interactive_player_performance_timeline.html')
        print("[ Time Slider Chart: No frames to animate; wrote placeholder HTML")
        return 0
    df_anim = pd.concat(cumulative_frames, ignore_index=True)

    # Create interactive line chart with time navigation (no animation for continuity)
    fig = px.line(
        df_anim,
        x='date',
        y='rating',
        color='player_name',
        line_group='player_name',
        markers=True,
        hover_name='player_name',
        hover_data={
            'date': False,  # use formatted date in customdata instead
            'rating': ':.2f'
        },
        custom_data=['team', 'acs', 'kills', 'deaths', 'assists', 'map_name', 'stage', 'date_display'],
        title='Top Players Performance Over Time - Valorant Champions 2024',
        labels={
            'rating': 'Player Rating',
            'date': 'Match Date',
            'player_name': 'Player'
        },
        animation_frame='frame',
        animation_group='player_name',
        category_orders={'frame': sorted(df_anim['frame'].unique())}
    )
    
    # Update layout for better appearance
    fig.update_layout(
        title={
            'text': 'Top Players Performance Over Time - Valorant Champions 2024',
            'x': 0.5,
            'xanchor': 'center',
            'font': {'size': 20, 'color': '#2E86AB'}
        },
        xaxis_title="Match Date",
        yaxis_title="Player Rating",
        width=1200,
        height=700,
        font=dict(size=12),
        plot_bgcolor='rgba(240,240,240,0.1)',
        paper_bgcolor='white',
        legend=dict(
            orientation="v",
            yanchor="top",
            y=1,
            xanchor="left",
            x=1.02
        )
    )
    
    # Update traces for better line appearance
    fig.update_traces(
        mode='lines+markers',
        line=dict(width=3),
        marker=dict(size=8, line=dict(width=2, color='white')),
        hovertemplate='<b>%{hovertext}</b><br>' +
                     'Team: %{customdata[0]}<br>' +
                     'Rating: %{y:.2f}<br>' +
                     'ACS: %{customdata[1]}<br>' +
                     'K/D/A: %{customdata[2]}/%{customdata[3]}/%{customdata[4]}<br>' +
                     'Map: %{customdata[5]}<br>' +
                     'Stage: %{customdata[6]}<br>' +
                     'Date: %{customdata[7]}<br>' +
                     '<extra></extra>'
    )

    # Ensure markers/hover settings apply to animation frames as well
    if fig.frames:
        for frame in fig.frames:
            for tr in frame.data:
                tr.update(
                    mode='lines+markers',
                    line=dict(width=3),
                    marker=dict(size=8, line=dict(width=2, color='white')),
                    hovertemplate='<b>%{hovertext}</b><br>' +
                                  'Team: %{customdata[0]}<br>' +
                                  'Rating: %{y:.2f}<br>' +
                                  'ACS: %{customdata[1]}<br>' +
                                  'K/D/A: %{customdata[2]}/%{customdata[3]}/%{customdata[4]}<br>' +
                                  'Map: %{customdata[5]}<br>' +
                                  'Stage: %{customdata[6]}<br>' +
                                  'Date: %{customdata[7]}<br>' +
                                  '<extra></extra>'
                )
    
    # Update x-axis
    fig.update_xaxes(
        showgrid=True,
        gridwidth=1,
        gridcolor='lightgray',
        tickangle=45
    )
    
    # Update y-axis
    fig.update_yaxes(
        showgrid=True,
        gridwidth=1,
        gridcolor='lightgray',
        range=[0.5, 1.5]  # Focus on relevant rating range
    )
    
    # Animation works best without range slider/selectors
    fig.update_layout(
        xaxis=dict(
            rangeslider=dict(visible=False),
            type="date"
        ),
        transition=dict(duration=300),
        updatemenus=[{
            'type': 'buttons',
            'showactive': True,
            'buttons': [
                {
                    'label': 'Play',
                    'method': 'animate',
                    'args': [None, {'frame': {'duration': 400, 'redraw': False}, 'fromcurrent': True, 'transition': {'duration': 300}}]
                },
                {
                    'label': 'Pause',
                    'method': 'animate',
                    'args': [[None], {'frame': {'duration': 0, 'redraw': False}, 'mode': 'immediate', 'transition': {'duration': 0}}]
                }
            ]
        }]
    )
    
    # Save as HTML
    fig.write_html('charts/interactive_player_performance_timeline.html')
    
    print(f"Time Slider Chart: Interactive player performance timeline - {len(df)} data points")
    return len(df)

def export_to_excel(conn):
    """Export data to Excel with formatting"""
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Sheet 1: Player Statistics
    ws1 = wb.create_sheet("Player Statistics")
    query1 = """
    SELECT 
        ps.player_name,
        ps.team,
        ps.rating,
        ps.acs,
        ps.kd_ratio,
        ps.rounds,
        ps.kills,
        ps.deaths,
        ps.assists
    FROM player_stats ps
    ORDER BY ps.rating DESC
    """
    df1 = execute_query(conn, query1, "Player Statistics")
    
    if df1 is not None and not df1.empty:
        # Add data to worksheet
        for r in dataframe_to_rows(df1, index=False, header=True):
            ws1.append(r)
        
        # Format headers
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Apply borders to all data cells
        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
            for cell in row:
                cell.border = border
        
        # Freeze headers
        ws1.freeze_panes = 'A2'
        
        # Add filters
        ws1.auto_filter.ref = ws1.dimensions
        
        # Apply conditional formatting for rating column (column C)
        rating_range = f"C2:C{ws1.max_row}"
        ws1.conditional_formatting.add(rating_range, 
            ColorScaleRule(start_type='min', start_color='FF6B6B',
                          end_type='max', end_color='4ECDC4'))
        
        # Apply conditional formatting for ACS column (column D)
        acs_range = f"D2:D{ws1.max_row}"
        ws1.conditional_formatting.add(acs_range,
            ColorScaleRule(start_type='min', start_color='FFE66D',
                          end_type='max', end_color='FF6B6B'))
    
    # Sheet 2: Team Performance
    ws2 = wb.create_sheet("Team Performance")
    query2 = """
    SELECT 
        ps.team,
        COUNT(*) as player_count,
        ROUND(AVG(ps.rating), 2) as avg_rating,
        ROUND(AVG(ps.acs), 2) as avg_acs,
        ROUND(AVG(ps.kd_ratio), 2) as avg_kd,
        ROUND(MIN(ps.rating), 2) as min_rating,
        ROUND(MAX(ps.rating), 2) as max_rating
    FROM player_stats ps
    GROUP BY ps.team
    ORDER BY avg_rating DESC
    """
    df2 = execute_query(conn, query2, "Team Performance")
    
    if df2 is not None and not df2.empty:
        for r in dataframe_to_rows(df2, index=False, header=True):
            ws2.append(r)
        
        # Format headers
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Apply borders
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
            for cell in row:
                cell.border = border
        
        # Freeze headers
        ws2.freeze_panes = 'A2'
        ws2.auto_filter.ref = ws2.dimensions
        
        # Apply conditional formatting for average rating
        rating_range = f"C2:C{ws2.max_row}"
        ws2.conditional_formatting.add(rating_range,
            ColorScaleRule(start_type='min', start_color='FF6B6B',
                          end_type='max', end_color='4ECDC4'))
    
    # Sheet 3: Map Statistics
    ws3 = wb.create_sheet("Map Statistics")
    query3 = """
    SELECT 
        ms.map_name,
        ms.times_played,
        ms.attack_win_percent,
        ms.defense_win_percent,
        CASE 
            WHEN ms.attack_win_percent > ms.defense_win_percent THEN 'Attack Favored'
            WHEN ms.defense_win_percent > ms.attack_win_percent THEN 'Defense Favored'
            ELSE 'Balanced'
        END as map_balance
    FROM maps_stats ms
    ORDER BY ms.times_played DESC
    """
    df3 = execute_query(conn, query3, "Map Statistics")
    
    if df3 is not None and not df3.empty:
        for r in dataframe_to_rows(df3, index=False, header=True):
            ws3.append(r)
        
        # Format headers
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Apply borders
        for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=ws3.max_column):
            for cell in row:
                cell.border = border
        
        # Freeze headers
        ws3.freeze_panes = 'A2'
        ws3.auto_filter.ref = ws3.dimensions
        
        # Apply conditional formatting for attack win percentage
        attack_range = f"C2:C{ws3.max_row}"
        ws3.conditional_formatting.add(attack_range,
            ColorScaleRule(start_type='min', start_color='FF6B6B',
                          end_type='max', end_color='4ECDC4'))
        
        # Apply conditional formatting for defense win percentage
        defense_range = f"D2:D{ws3.max_row}"
        ws3.conditional_formatting.add(defense_range,
            ColorScaleRule(start_type='min', start_color='FF6B6B',
                          end_type='max', end_color='4ECDC4'))
    
    # Save workbook
    filename = f'exports/valorant_champions_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(filename)
    
    # Calculate total rows
    total_rows = 0
    if df1 is not None:
        total_rows += len(df1)
    if df2 is not None:
        total_rows += len(df2)
    if df3 is not None:
        total_rows += len(df3)
    
    print(f"Excel Export: Created {filename}")
    print(f"  - 3 sheets: Player Statistics, Team Performance, Map Statistics")
    print(f"  - Total rows: {total_rows}")
    print(f"  - Features: Frozen headers, gradient fills, filters, conditional formatting")
    
    return filename

def main():
    """Main function to create all visualizations and exports"""
    print("VALORANT CHAMPIONS 2024 - DATA VISUALIZATION & EXPORT")
    print("=" * 60)
    
    # Create directories
    create_charts_directory()
    
    # Connect to database
    conn = connect_to_db()
    if not conn:
        return
    
    try:
        # Create all charts
        print("\nCreating Visualizations...")
        print("-" * 30)
        
        pie_rows = create_pie_chart(conn)
        bar_rows = create_bar_chart(conn)
        hbar_rows = create_horizontal_bar_chart(conn)
        line_rows = create_line_chart(conn)
        hist_rows = create_histogram(conn)
        scatter_rows = create_scatter_plot(conn)
        timeline_rows = create_time_slider_chart(conn)
        
        # Export to Excel
        print("\nCreating Excel Export...")
        print("-" * 30)
        excel_file = export_to_excel(conn)
        
        # Print summary report
        print(f"\n{'='*60}")
        print("VISUALIZATION SUMMARY REPORT")
        print(f"{'='*60}")
        print(f"Pie Chart: Team rating distribution - {pie_rows} teams")
        print(f"Bar Chart: Top players by ACS - {bar_rows} players")
        print(f"Horizontal Bar: Map win rates - {hbar_rows} maps")
        print(f"Line Chart: Player performance over time - {line_rows} data points")
        print(f"Histogram: Player ratings distribution - {hist_rows} players")
        print(f"Scatter Plot: ACS vs Rating correlation - {scatter_rows} players")
        print(f"Time Slider: Interactive player performance timeline - {timeline_rows} data points")
        print(f"Excel Export: {excel_file}")
        print(f"\nAll charts saved to /charts/ directory")
        print(f"Excel report saved to /exports/ directory")
        print(f"{'='*60}")
        
    except Exception as e:
        print(f"Error during visualization creation: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        conn.close()
        print("\nDatabase connection closed.")

if __name__ == "__main__":
    main()
